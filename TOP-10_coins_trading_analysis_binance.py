
# C:\Work\Blockchain\MT_Михаил_Трофимов\ML\bybit\bbt_004\TOP-10.py

import os
import pandas as pd
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Функция для поиска файла CSV в текущей директории
def find_csv_file():
    for file in os.listdir():
        if file.endswith('.csv'):
            return file
    raise FileNotFoundError("CSV файл не найден в текущей директории.")

# Автоматически найти и использовать CSV файл
file_path = find_csv_file()
print(f"Используется файл: {file_path}")


df = pd.read_csv(file_path)

# Удаление пробелов из имен столбцов
df.columns = df.columns.str.strip()

# Удаление пробелов в значениях столбца 'Date/Time'
df['Date/Time'] = df['Date/Time'].str.strip()

# Преобразование столбца 'Date/Time' в формат datetime
df['Date/Time'] = pd.to_datetime(df['Date/Time'], format='%y/%m/%d %H:%M:%S', errors='coerce')

# Удаление строк с некорректными датами
df = df.dropna(subset=['Date/Time'])

# Преобразование всех символов в столбце 'Symbol' в нижний регистр
df['Symbol'] = df['Symbol'].str.lower()

# Группировка данных по символу монеты и суммирование дохода с учётом знака
grouped_df = df.groupby('Symbol')['PNLUSDT'].sum()

# Сортировка по убыванию дохода
sorted_df = grouped_df.sort_values(ascending=False)

# Получение топ-10 монет
top_10 = sorted_df.head(10)

# Фильтрация исходного DataFrame для топ-10 монет
top_10_symbols = top_10.index
top_10_df = df[df['Symbol'].isin(top_10_symbols)]

# Группировка данных по 'Symbol' и 'Opened' и вычисление дохода с учётом знака
grouped_by_symbol_and_algo = top_10_df.groupby(['Symbol', 'Opened'])['PNLUSDT'].sum().reset_index()

# Нахождение алгоритма с максимальным доходом для каждой монеты
max_algo_for_each_symbol = grouped_by_symbol_and_algo.loc[grouped_by_symbol_and_algo.groupby('Symbol')['PNLUSDT'].idxmax()]

# Создание результирующей таблицы
result_table = max_algo_for_each_symbol[['Symbol', 'Opened', 'PNLUSDT']]

# Переименование столбцов для удобства
result_table.columns = ['Symbol', 'Algorithm', 'Max PNLUSDT']

# Фильтрация сделок за последние 30 дней
last_30_days = df[df['Date/Time'] >= (pd.Timestamp.now() - pd.Timedelta(days=30))]

# Подсчет количества сделок за последние 30 дней для каждой монеты
trades_last_30_days = last_30_days['Symbol'].value_counts().reindex(top_10_symbols).fillna(0).astype(int).reset_index()
trades_last_30_days.columns = ['Symbol', 'Trades last 30 days']

# Подсчет только плюсовых и минусовых сделок за последние 30 дней
positive_trades = last_30_days[last_30_days['PNLUSDT'] > 0]
negative_trades = last_30_days[last_30_days['PNLUSDT'] < 0]

positive_trades_count = positive_trades['Symbol'].value_counts().reindex(top_10_symbols).fillna(0).astype(int).reset_index()
positive_trades_count.columns = ['Symbol', 'Positive Trades Count']

negative_trades_count = negative_trades['Symbol'].value_counts().reindex(top_10_symbols).fillna(0).astype(int).reset_index()
negative_trades_count.columns = ['Symbol', 'Negative Trades Count']

positive_trades_sum = positive_trades.groupby('Symbol')['PNLUSDT'].sum().reindex(top_10_symbols).fillna(0).reset_index()
positive_trades_sum.columns = ['Symbol', 'Positive Trades Sum']

negative_trades_sum = negative_trades.groupby('Symbol')['PNLUSDT'].sum().reindex(top_10_symbols).fillna(0).reset_index()
negative_trades_sum.columns = ['Symbol', 'Negative Trades Sum']

# Объединение данных
final_result = result_table.merge(trades_last_30_days, on='Symbol')
final_result = final_result.merge(positive_trades_sum, on='Symbol')
final_result = final_result.merge(positive_trades_count, on='Symbol')
final_result = final_result.merge(negative_trades_sum, on='Symbol')
final_result = final_result.merge(negative_trades_count, on='Symbol')

# Пересчет Max PNLUSDT
final_result['Max PNLUSDT'] = final_result['Positive Trades Sum'] + final_result['Negative Trades Sum']

# Проверка наличия столбца 'Max PNLUSDT' в final_result перед суммированием
if 'Max PNLUSDT' in final_result.columns:
    total_pnl = final_result['Max PNLUSDT'].sum()
    print(f"\nTOTAL: {total_pnl:.2f} USDT")
else:
    print("Ошибка: Столбец 'Max PNLUSDT' не найден в результирующей таблице.")

# Вывод обновленной таблицы
print("Updated Top 10 Trades Count:")
print(final_result)

# таблица для монет с максимальным убытком (минус топ-10)
min_10 = sorted_df.tail(10)
min_10_symbols = min_10.index
min_10_df = df[df['Symbol'].isin(min_10_symbols)]

# Группировка данных по 'Symbol' и 'Opened' и вычисление убытков
grouped_by_symbol_and_algo_min = min_10_df.groupby(['Symbol', 'Opened'])['PNLUSDT'].sum().reset_index()

# Нахождение алгоритма с максимальным убытком для каждой монеты
min_algo_for_each_symbol = grouped_by_symbol_and_algo_min.loc[grouped_by_symbol_and_algo_min.groupby('Symbol')['PNLUSDT'].idxmin()]

# Создание результирующей таблицы для убытков
result_table_min = min_algo_for_each_symbol[['Symbol', 'Opened', 'PNLUSDT']]

# Переименование столбцов для удобства
result_table_min.columns = ['Symbol', 'Algorithm', 'Min PNLUSDT']

# Подсчет количества сделок за последние 30 дней для минус топ-10 монет
trades_last_30_days_min = last_30_days['Symbol'].value_counts().reindex(min_10_symbols).fillna(0).astype(int).reset_index()
trades_last_30_days_min.columns = ['Symbol', 'Trades last 30 days']

# Подсчет только плюсовых и минусовых сделок за последние 30 дней для минус топ-10 монет
positive_trades_min = last_30_days[last_30_days['PNLUSDT'] > 0]
negative_trades_min = last_30_days[last_30_days['PNLUSDT'] < 0]

positive_trades_count_min = positive_trades_min['Symbol'].value_counts().reindex(min_10_symbols).fillna(0).astype(int).reset_index()
positive_trades_count_min.columns = ['Symbol', 'Positive Trades Count']

negative_trades_count_min = negative_trades_min['Symbol'].value_counts().reindex(min_10_symbols).fillna(0).astype(int).reset_index()
negative_trades_count_min.columns = ['Symbol', 'Negative Trades Count']

positive_trades_sum_min = positive_trades_min.groupby('Symbol')['PNLUSDT'].sum().reindex(min_10_symbols).fillna(0).reset_index()
positive_trades_sum_min.columns = ['Symbol', 'Positive Trades Sum']

negative_trades_sum_min = negative_trades_min.groupby('Symbol')['PNLUSDT'].sum().reindex(min_10_symbols).fillna(0).reset_index()
negative_trades_sum_min.columns = ['Symbol', 'Negative Trades Sum']

# Объединение данных для минус топ-10 монет
final_result_min = result_table_min.merge(trades_last_30_days_min, on='Symbol')
final_result_min = final_result_min.merge(positive_trades_sum_min, on='Symbol')
final_result_min = final_result_min.merge(positive_trades_count_min, on='Symbol')
final_result_min = final_result_min.merge(negative_trades_sum_min, on='Symbol')
final_result_min = final_result_min.merge(negative_trades_count_min, on='Symbol')

# Пересчет Min PNLUSDT
final_result_min['Min PNLUSDT'] = final_result_min['Positive Trades Sum'] + final_result_min['Negative Trades Sum']

# Проверка наличия столбца 'Min PNLUSDT' в final_result_min перед суммированием
if 'Min PNLUSDT' in final_result_min.columns:
    total_pnl_min = final_result_min['Min PNLUSDT'].sum()
    print(f"\nTOTAL (negative): {total_pnl_min:.2f} USDT")
else:
    print("Ошибка: Столбец 'Min PNLUSDT' не найден в результирующей таблице.")

# Вывод таблицы для минус топ-10 монет
print("Updated Minus Top 10 Trades Count:")
print(final_result_min)

# Запись данных в Excel
with pd.ExcelWriter('result.xlsx') as writer:
    final_result.to_excel(writer, sheet_name='Top 10 Trades Count', index=False)
    final_result_min.to_excel(writer, sheet_name='Minus Top 10 Trades Count', index=False)

# Открытие созданного Excel файла и добавление фильтров и закрепления строк
wb = load_workbook('result.xlsx')
ws_top_10 = wb['Top 10 Trades Count']
ws_min_10 = wb['Minus Top 10 Trades Count']

# Добавление фильтра в первую строку для обеих таблиц
ws_top_10.auto_filter.ref = ws_top_10.dimensions
ws_min_10.auto_filter.ref = ws_min_10.dimensions

# Закрепление верхней строки (Freeze Top Row) для обеих таблиц
ws_top_10.freeze_panes = ws_top_10['A2']
ws_min_10.freeze_panes = ws_min_10['A2']

# Сохранение изменений
wb.save('result.xlsx')
